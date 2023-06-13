from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("Student Registration System")
root.geometry("1050x710")
root.config(bg=background)
root.resizable(False,False)




###########ShowImage##########
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetypes=(("JPG File","*.jpg"),
                                                                             ("PNG File","*.png"),
                                                                             ("All files","*.txt")))
    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2


##########Clear###########
def Clear():
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")

    registration_no()

    saveButton.config(state='normal')

    img1=PhotoImage(file='Images/upload photo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""

##########Save###########
def Save():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=Re1)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)

        file.save(r'Student_data.xlsx')

        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not available!!!!")

        messagebox.showinfo("info","Successfully data entered!!!")

        Clear()    #clear entry box and image section

        registration_no()  #it will recheck registration no. and reissue new no.

###############Search##########
def search():
    text=Search.get()  #taking input from entry box

    Clear()  # to clear all the data already available in entry box and other
    saveButton.config(state='disable')

    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number!!!")

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)
    
    if x4=="Female":
        R2.select()
    elif x4=="Male":
        R1.select()
    else:
        R3.select()

    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    img=(Image.open("Student Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#############Update##############
def Update():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    selection()
    G1=gender
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active

    for row in sheet.rows():
        if row[0].value==R1:
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    #sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=C1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=Re1)
    sheet.cell(column=8,row=int(reg_number),value=S1)
    sheet.cell(column=9,row=int(reg_number),value=fathername)
    sheet.cell(column=10,row=int(reg_number),value=mothername)
    sheet.cell(column=11,row=int(reg_number),value=F1)
    sheet.cell(column=12,row=int(reg_number),value=M1)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update","Update Successfully!!")

    Clear()

#############Delete###########
def delete():

    R1=Registration.get()
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active

    sheet.delete_rows(idx=R1)

    file.save(r'Student_data.xlsx')

    messagebox.showinfo("Delete","Student data deleted successfully!!!")

    Clear()

#gender
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
    elif value==2:
        gender="Female"
    else:
        gender="Others"


#Student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=990,bg=framebg,fg=framefg,height=650,relief=GROOVE)
obj.place(x=30,y=30)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=400,y=50)
Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=400,y=100)
Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=400,y=150)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)

radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)

R3=Radiobutton(obj,text="Others",variable=radio,value=3,bg=framebg,fg=framefg,command=selection)
R3.place(x=260,y=150)

Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,width=20,font="arial 10")
religion_entry.place(x=530,y=100)

Skill=StringVar()
skill_entry=Entry(obj,textvariable=Skill,width=20,font="arial 10")
skill_entry.place(x=530,y=150)

Class=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 10",width=17,state="r")
Class.place(x=530,y=50)
Class.set("Select Class")


#image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=800,y=80)

img=PhotoImage(file="Images/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#button
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=800,y=300)

root.mainloop()